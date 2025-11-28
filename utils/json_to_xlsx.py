# -*- coding: utf-8 -*-
"""
JSON to XLSX 변환 유틸리티
"""
import re
import pandas as pd
from pathlib import Path
from tinydb import TinyDB
from tinydb.storages import JSONStorage
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from .print_log import print


class UTF8JSONStorage(JSONStorage):
    """UTF-8 인코딩을 지원하는 JSON 스토리지"""
    
    def __init__(self, path, **kwargs):
        super().__init__(path, encoding='utf-8', **kwargs)


def _make_table_name(base: str, index: int) -> str:
    """Excel 테이블 DisplayName을 생성합니다."""
    safe = re.sub(r'[^0-9A-Za-z_]', '_', base)
    if not safe or safe[0].isdigit():
        safe = f"tbl_{safe}"
    # Excel 테이블 이름은 31자 제한 권장
    safe = safe[:25]
    return f"{safe}_{index}"


def json_to_xlsx(db_path: Path):
    """
    TinyDB JSON 파일을 XLSX 파일로 변환합니다.
    
    Args:
        db_path: 데이터베이스 파일 경로
    """
    if not db_path.exists():
        print.Warn(f"데이터베이스 파일이 없습니다: {db_path}")
        return
    
    db = TinyDB(db_path, storage=UTF8JSONStorage)
    table_names = db.tables()
    
    if not table_names:
        print.Warn("테이블이 없습니다")
        return
    
    new_file = db_path.with_suffix('.xlsx')
    if new_file.exists():
        new_file.unlink()
    
    try:
        with pd.ExcelWriter(new_file, engine='openpyxl') as writer:
            for idx, table_name in enumerate(table_names):
                table = db.table(table_name)
                data = table.all()
                
                # 리스트를 문자열로 변환
                for row in data:
                    for col, value in row.items():
                        if isinstance(value, list):
                            row[col] = ', '.join(map(str, value))
                
                df = pd.DataFrame(data)

                if 'count' in df.columns:
                    df['count'] = pd.to_numeric(df['count'], errors='coerce')
                    df = df.sort_values(by='count', ascending=False, na_position='last')
                sheet_name = str(table_name)[:31]  # Excel 시트 이름 제한
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]

                # 표 스타일 적용 (데이터가 있을 때만)
                if not df.empty:
                    last_col = get_column_letter(len(df.columns))
                    last_row = len(df.index) + 1  # 헤더 포함
                    table_ref = f"A1:{last_col}{last_row}"
                    display_name = _make_table_name(table_name, idx)
                    xl_table = Table(displayName=display_name, ref=table_ref)
                    style = TableStyleInfo(
                        name="TableStyleMedium9",
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=False,
                    )
                    xl_table.tableStyleInfo = style
                    worksheet.add_table(xl_table)

                # 열 너비 자동 조정
                for col_idx, col in enumerate(df.columns, start=1):
                    col_letter = get_column_letter(col_idx)
                    series_max = df[col].astype(str).map(len).max()
                    if pd.isna(series_max):
                        series_max = 0
                    max_len = max(series_max, len(str(col)))
                    max_len = min(max_len, 200)
                    worksheet.column_dimensions[col_letter].width = max_len + 2
        
        print.Info("XLSX 파일 생성 완료:", new_file)
    except Exception as e:
        print.exception(show_locals=True)

